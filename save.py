import openpyxl as op


name = '0432354'

wb = op.load_workbook('madreseye taht app.xlsx')
sheet = wb['Classes']

if sheet.cell(row=1, column=1).value == None:
	sheet.cell(row=1, column=1).value = name
else:
	sheet.cell(row=1, column=sheet.max_column+1).value = name

wb.save('madreseye taht app.xlsx')