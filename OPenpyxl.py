import openpyxl
#baraye baz kardan yek safhe
wb = openpyxl.load_rorkbook('file path')

#baraye fahmidan esm haye sheet haye filemon
wb.get_sheet_names()

#braye datresy dashtan be safheye mored nazar
sheet = wb.get_sheet_by_name('sheet name')

#braye inke be yek motaghayer datressy peyda koni
sheet['B2'].value
#ya
sheet.cell(row=1, column=3).value

#barayeanke motavageh shim chand satr vogod dare
sheet.max_row
#baraye ston ha
sheet.max_column


#baraye anke befahmim radif 4 dar excel che harfie
openpyxl.cell.get_column_letter(4)
#baraks
openpyxl.cel.column_index_from_string('D')


#braye inke title ye safhe ro befahmim
sheet.title

#dorostkardan
#safhe
wb.create_sheet(title='', index='')
#excel
wb = openpyxl.Workbook()
